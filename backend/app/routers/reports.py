import csv
import io
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import Optional

from app.core.database import get_db
from app.core.deps import require_manager_or_admin
from app.models.user import User
from app.models.goal import Goal, GoalStatus
from app.models.achievement import Achievement
from app.models.goal_cycle import GoalCycle

router = APIRouter(prefix="/reports", tags=["reports"])


def _get_report_data(db: Session, current_user: User,
                      cycle_id: Optional[int] = None,
                      department_id: Optional[int] = None,
                      status: Optional[str] = None):
    query = (
        db.query(Achievement, Goal, User)
        .join(Goal, Achievement.goal_id == Goal.id)
        .join(User, Goal.employee_id == User.id)
        .options(joinedload(Goal.thrust_area))
    )

    if current_user.role == "manager":
        team_ids = [u.id for u in db.query(User).filter(
            User.manager_id == current_user.id).all()]
        query = query.filter(Goal.employee_id.in_(team_ids))

    if cycle_id:
        query = query.filter(Achievement.cycle_id == cycle_id)
    if department_id:
        query = query.filter(User.department_id == department_id)
    if status:
        query = query.filter(Achievement.goal_status == status)

    return query.all()


@router.get("/achievement/csv")
def export_achievement_csv(
    cycle_id: Optional[int] = None,
    department_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_or_admin),
):
    rows = _get_report_data(db, current_user, cycle_id, department_id, status)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Employee Name", "Goal Title", "Thrust Area", "UoM Type",
        "Target", "Actual Achievement", "Tracking Score", "Status", "Cycle"
    ])

    for ach, goal, user in rows:
        cycle = db.query(GoalCycle).filter(GoalCycle.id == ach.cycle_id).first()
        writer.writerow([
            user.name,
            goal.title,
            goal.thrust_area.name if goal.thrust_area else "",
            goal.uom_type,
            goal.target or goal.target_date,
            ach.actual_value or ach.actual_date,
            f"{ach.tracking_score:.1f}%" if ach.tracking_score is not None else "N/A",
            ach.goal_status,
            cycle.name if cycle else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=achievement_report.csv"},
    )


@router.get("/achievement/excel")
def export_achievement_excel(
    cycle_id: Optional[int] = None,
    department_id: Optional[int] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_manager_or_admin),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = _get_report_data(db, current_user, cycle_id, department_id, status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Achievement Report"

    headers = [
        "Employee Name", "Goal Title", "Thrust Area", "UoM Type",
        "Target", "Actual Achievement", "Tracking Score", "Status", "Cycle"
    ]

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (ach, goal, user) in enumerate(rows, 2):
        cycle = db.query(GoalCycle).filter(GoalCycle.id == ach.cycle_id).first()
        ws.cell(row=row_idx, column=1, value=user.name)
        ws.cell(row=row_idx, column=2, value=goal.title)
        ws.cell(row=row_idx, column=3, value=goal.thrust_area.name if goal.thrust_area else "")
        ws.cell(row=row_idx, column=4, value=goal.uom_type)
        ws.cell(row=row_idx, column=5, value=str(goal.target or goal.target_date))
        ws.cell(row=row_idx, column=6, value=str(ach.actual_value or ach.actual_date or ""))
        ws.cell(row=row_idx, column=7,
                value=f"{ach.tracking_score:.1f}%" if ach.tracking_score is not None else "N/A")
        ws.cell(row=row_idx, column=8, value=ach.goal_status)
        ws.cell(row=row_idx, column=9, value=cycle.name if cycle else "")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=achievement_report.xlsx"},
    )
